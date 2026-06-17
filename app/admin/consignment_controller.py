import io
import logging
import mimetypes
import re
import os
import uuid
import base64
import binascii
import tempfile
from datetime import datetime
from flask import current_app
from werkzeug.utils import secure_filename
import io as _io

# Supabase integration is optional: when SUPABASE_URL and SUPABASE_KEY are set
# and the `supabase` package is available, uploads will go to Supabase Storage.
def _get_supabase_client():
    url = os.getenv('SUPABASE_URL', '').strip()
    key = os.getenv('SUPABASE_KEY', '').strip()
    if not url or not key:
        return None
    try:
        from supabase import create_client
    except Exception:
        logger.warning('Supabase package not available; falling back to local uploads')
        return None

    try:
        return create_client(url, key)
    except Exception:
        logger.exception('Failed to create Supabase client')
        return None


def _decode_pod_data_url(data_url):
    if not data_url or not isinstance(data_url, str):
        raise ValueError('POD file data is missing.')
    if ',' not in data_url:
        raise ValueError('Invalid POD file data.')

    header, encoded = data_url.split(',', 1)
    if not header.startswith('data:image/') or ';base64' not in header:
        raise ValueError('POD file data must be a base64 encoded image.')

    try:
        file_bytes = base64.b64decode(encoded, validate=True)
    except (binascii.Error, ValueError):
        raise ValueError('POD file data is invalid.')

    if len(file_bytes) > MAX_POD_IMAGE_BYTES:
        raise ValueError('POD image must be smaller than 5 MB.')

    return file_bytes


def _store_pod_bytes(filename, file_bytes, content_type=None, bucket_name=None):
    supa = _get_supabase_client()
    if supa:
        bucket = bucket_name or os.getenv('SUPABASE_BUCKET', 'pod-uploads')
        object_path = f"consignments/{filename}"
        # storage3 client expects a file path; write bytes to a temporary file
        tmp = None
        try:
            tf = tempfile.NamedTemporaryFile(delete=False)
            tmp = tf.name
            tf.write(file_bytes)
            tf.flush()
            tf.close()
            supa.storage.from_(bucket).upload(
                object_path,
                tmp,
                {'content-type': content_type or 'application/octet-stream'},
            )
        finally:
            try:
                if tmp and os.path.exists(tmp):
                    os.remove(tmp)
            except Exception:
                logger.exception('Failed to remove temporary POD upload file')
        return f"supabase:{bucket}/{object_path}"

    upload_folder = os.path.join(current_app.instance_path, "uploads")
    os.makedirs(upload_folder, exist_ok=True)
    dest_path = os.path.join(upload_folder, filename)
    with open(dest_path, "wb") as file_handle:
        file_handle.write(file_bytes)
    return filename


def _parse_supabase_pod_value(pod_value):
    if not isinstance(pod_value, str) or not pod_value.startswith('supabase:'):
        raise ValueError('POD is not stored in Supabase.')

    _, rest = pod_value.split(':', 1)
    bucket, object_path = rest.split('/', 1)
    if not bucket or not object_path:
        raise ValueError('Invalid Supabase POD path.')

    return bucket, object_path


def _download_supabase_pod_file(pod_value):
    client = _get_supabase_client()
    if not client:
        raise RuntimeError('Supabase not configured.')

    bucket, object_path = _parse_supabase_pod_value(pod_value)
    content = client.storage.from_(bucket).download(object_path)
    if hasattr(content, 'read'):
        content = content.read()
    if isinstance(content, bytearray):
        content = bytes(content)
    if not isinstance(content, bytes):
        raise RuntimeError('Unexpected Supabase download response.')

    return content, object_path


def _download_legacy_supabase_pod_file(consignment_id, pod_value):
    """Download a legacy POD by attempting old Supabase object paths.

    Legacy records may store a bare object path or a local filename that was
    previously migrated into Supabase. This helper tries the configured bucket
    and an optional consignment-id-prefixed path.
    """
    client = _get_supabase_client()
    if not client:
        raise RuntimeError('Supabase not configured.')

    bucket = os.getenv('SUPABASE_BUCKET', 'pod-uploads')
    if not isinstance(pod_value, str) or not pod_value:
        raise ValueError('Invalid legacy POD path.')

    candidates = []
    legacy_bucket = bucket
    if pod_value.startswith('supabase:'):
        _, rest = pod_value.split(':', 1)
        try:
            legacy_bucket, legacy_object_path = rest.split('/', 1)
        except ValueError:
            legacy_bucket = bucket
            legacy_object_path = rest
        candidates.append((legacy_bucket, legacy_object_path))
    else:
        candidates.append((bucket, pod_value))
        if consignment_id is not None:
            consignment_prefix = str(consignment_id)
            if not pod_value.startswith(consignment_prefix + '/'):
                candidates.append((bucket, f"{consignment_prefix}/{pod_value}"))

    last_error = None
    for candidate_bucket, object_path in candidates:
        try:
            content = client.storage.from_(candidate_bucket).download(object_path)
            if hasattr(content, 'read'):
                content = content.read()
            if isinstance(content, bytearray):
                content = bytes(content)
            if not isinstance(content, bytes):
                raise RuntimeError('Unexpected Supabase download response.')
            return content, candidate_bucket, object_path
        except Exception as exc:
            last_error = exc

    # Last chance: if the value was a local filename under uploads, try that path.
    upload_folder = os.path.join(current_app.instance_path, 'uploads')
    try:
        legacy_path = os.path.normpath(os.path.join(upload_folder, pod_value))
        if legacy_path.startswith(os.path.abspath(upload_folder)) and os.path.exists(legacy_path):
            with open(legacy_path, 'rb') as fh:
                return fh.read(), bucket, pod_value
    except Exception:
        pass

    raise RuntimeError('Legacy POD file not found.') from last_error


def _delete_pod_file(pod_value):
    if not pod_value:
        return

    if isinstance(pod_value, str) and pod_value.startswith('supabase:'):
        client = _get_supabase_client()
        if not client:
            return
        try:
            _, rest = pod_value.split(':', 1)
            bucket, object_path = rest.split('/', 1)
            client.storage.from_(bucket).remove([object_path])
        except Exception:
            logger.exception('Failed to remove POD from Supabase')
        return

    upload_folder = os.path.join(current_app.instance_path, "uploads")
    pod_path = os.path.normpath(os.path.join(upload_folder, pod_value))
    if pod_path.startswith(os.path.abspath(upload_folder)) and os.path.exists(pod_path):
        try:
            os.remove(pod_path)
        except Exception:
            logger.exception('Failed to remove POD file from disk')


def _parse_date_string(value):
    if not value or not isinstance(value, str):
        return None

    value = value.strip()
    if not value:
        return None

    # Prefer ISO date format, but support common day/month/year formats if present.
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue

    return None

from flask import flash, jsonify, redirect, render_template, request, send_file, url_for
from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.exc import DatabaseError, IntegrityError, OperationalError, ProgrammingError
from sqlalchemy import inspect as sa_inspect

from app import limiter
from app.admin import admin_bp
from app.admin.auth import require_admin
from app.models import Consignment, db
from app.services.logistics import (
    normalize_consignment_number,
    normalize_indian_pincode,
    normalize_status,
)

logger = logging.getLogger(__name__)

MAX_POD_IMAGE_BYTES = 5 * 1024 * 1024


def _is_external_pod_url(value):
    if not isinstance(value, str):
        return False
    lowered = value.strip().lower()
    return lowered.startswith("http://") or lowered.startswith("https://")


def _is_missing_column_error(error):
    """Return True when the database rejected a query because a column is absent."""
    original = getattr(error, "orig", None)
    if original is None:
        return False

    pgcode = getattr(original, "pgcode", None)
    if pgcode == "42703":
        return True

    return original.__class__.__name__ == "UndefinedColumn"


@admin_bp.route("/admin/consignments", methods=["GET"], endpoint="consignments_panel")
@require_admin
def consignments_panel():
    try:
        # Avoid loading the entire table into memory for the admin panel.
        # For large datasets render a small sample and let the client use the paginated API.
        total = Consignment.query.count()
        if total > 500:
            consignments = []
            logger.info("consignments_panel: large table detected (total=%d); rendering empty sample and deferring to API", total)
        else:
            consignments = Consignment.query.order_by(Consignment.id.asc()).limit(200).all()

        rows = [
            {
                "id": c.id,
                "consignment_number": c.consignment_number,
                "status": c.status,
                "pickup_pincode": c.pickup_pincode,
                "pickup_address": getattr(c, "pickup_address", None),
                "pickup_tag": getattr(c, "pickup_tag", None),
                "pickup_date": getattr(c, "pickup_date", None),
                "drop_pincode": c.drop_pincode,
                "drop_address": getattr(c, "drop_address", None),
                "drop_tag": getattr(c, "drop_tag", None),
                "drop_date": getattr(c, "drop_date", None),
                "eta": c.eta,
                "pod_image": getattr(c, "pod_image", None),
            }
            for c in consignments
        ]
        return render_template(
            "admin/consignments.html",
            consignments=rows,
        )
    except ProgrammingError as error:
        db.session.rollback()
        if _is_missing_column_error(error):
            logger.exception("Schema mismatch loading admin panel")
            return render_template(
                "admin/consignments.html",
                consignments=[],
                error="Database schema needs an update. Missing consignment fields.",
            )

        logger.exception("Database error loading admin panel")
        return render_template(
            "admin/consignments.html",
            consignments=[],
            error="Unable to load data. Please try again.",
        )
    except (OperationalError, DatabaseError):
        logger.exception("Database error loading admin panel")
        return render_template(
            "admin/consignments.html",
            consignments=[],
            error="Unable to load data. Please try again.",
        )
    except Exception:
        logger.exception("Unexpected error in admin panel")
        return render_template(
            "admin/consignments.html",
            consignments=[],
            error="An unexpected error occurred.",
        )


def _normalize_header(value):
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


@admin_bp.route("/admin/consignments/list", methods=["GET"], endpoint="consignments_list_api")
@require_admin
def consignments_list_api():
    """API endpoint for paginated, searchable, and sortable consignments."""
    try:
        # Get parameters from request
        page = request.args.get("page", 1, type=int)
        per_page = request.args.get("per_page", 10, type=int)
        search = request.args.get("search", "", type=str).strip()
        sort_by = request.args.get("sort_by", "id", type=str)
        sort_order = request.args.get("sort_order", "asc", type=str)

        # Validate pagination parameters
        page = max(1, page)
        per_page = max(1, min(100, per_page))  # Limit to max 100 per page

        # Validate sort parameters
        allowed_sort_columns = {
            "id", "consignment_number", "status", "pickup_pincode",
            "drop_pincode", "pickup_tag", "drop_tag", "pickup_date", "drop_date"
        }
        sort_by = sort_by if sort_by in allowed_sort_columns else "id"
        sort_order = "asc" if sort_order.lower() == "asc" else "desc"

        # Build base query
        query = Consignment.query

        # Apply search filter if provided
        if search:
            search_pattern = f"%{search}%"
            query = query.filter(
                db.or_(
                    Consignment.consignment_number.ilike(search_pattern),
                    Consignment.status.ilike(search_pattern),
                    Consignment.pickup_tag.ilike(search_pattern),
                    Consignment.drop_tag.ilike(search_pattern),
                    Consignment.pickup_pincode.ilike(search_pattern),
                    Consignment.drop_pincode.ilike(search_pattern),
                    Consignment.pickup_address.ilike(search_pattern),
                    Consignment.drop_address.ilike(search_pattern),
                )
            )

        # Get total count before pagination
        total = query.count()

        # Apply sorting
        sort_column = getattr(Consignment, sort_by)
        if sort_order == "desc":
            query = query.order_by(sort_column.desc())
        else:
            query = query.order_by(sort_column.asc())

        # Apply pagination
        paginated = query.paginate(page=page, per_page=per_page, error_out=False)

        # Serialize results
        rows = [
            {
                "id": c.id,
                "consignment_number": c.consignment_number,
                "status": c.status,
                "pickup_pincode": c.pickup_pincode,
                "pickup_address": getattr(c, "pickup_address", None),
                "pickup_tag": getattr(c, "pickup_tag", None),
                "pickup_date": getattr(c, "pickup_date", None),
                "drop_pincode": c.drop_pincode,
                "drop_address": getattr(c, "drop_address", None),
                "drop_tag": getattr(c, "drop_tag", None),
                "drop_date": getattr(c, "drop_date", None),
                "eta": c.eta,
                "pod_image": getattr(c, "pod_image", None),
            }
            for c in paginated.items
        ]

        return jsonify({
            "success": True,
            "rows": rows,
            "page": page,
            "per_page": per_page,
            "total": total,
            "pages": paginated.pages,
            "has_prev": paginated.has_prev,
            "has_next": paginated.has_next,
        })

    except ProgrammingError as error:
        db.session.rollback()
        if _is_missing_column_error(error):
            logger.exception("Schema mismatch in consignments API")
            return jsonify({
                "success": False,
                "error": "Database schema needs an update. Missing consignment fields."
            }), 500

        logger.exception("Database error in consignments API")
        return jsonify({
            "success": False,
            "error": "Unable to load data. Please try again."
        }), 500
    except (OperationalError, DatabaseError):
        db.session.rollback()
        logger.exception("Database error in consignments API")
        return jsonify({
            "success": False,
            "error": "Database connection error. Please try again."
        }), 500
    except Exception as e:
        logger.exception("Unexpected error in consignments API")
        return jsonify({
            "success": False,
            "error": "An unexpected error occurred."
        }), 500


@admin_bp.route("/admin/consignments/save", methods=["POST"], endpoint="consignments_save")
@limiter.limit("25 per minute")
@require_admin
def consignments_save():
    payload = request.get_json(silent=True) or {}
    rows = payload.get("rows", [])
    deleted_ids = payload.get("deleted_ids", [])

    if not isinstance(rows, list) or not isinstance(deleted_ids, list):
        return jsonify({"success": False, "message": "Invalid request payload."}), 400

    try:
        # Debug logging: record incoming payload and per-row id types to diagnose
        logger.info("consignments_save incoming payload: %s", payload)
        logger.info("consignments_save deleted_ids (raw): %s", deleted_ids)
        for idx, r in enumerate(rows):
            try:
                raw_id = r.get("id")
            except Exception:
                raw_id = None
            logger.info(
                "consignments_save row %d raw_id=%r type=%s consignment_number=%r",
                idx,
                raw_id,
                type(raw_id).__name__ if raw_id is not None else "None",
                r.get("consignment_number"),
            )

        # Avoid selecting all mapped columns (which fails if the DB schema is missing new columns).
        # Fetch only ids first so we can detect missing-column errors early and provide a clear message.
        try:
            existing_ids = [r[0] for r in db.session.query(Consignment.id).all()]
            existing = {int(i): None for i in existing_ids}
        except (ProgrammingError, OperationalError, DatabaseError) as e:
            db.session.rollback()
            if _is_missing_column_error(e):
                logger.exception("Schema mismatch in admin save")
                return jsonify({"success": False, "message": "Database schema needs an update. Missing consignment fields."}), 500

            logger.exception("Database error in admin save")
            return jsonify({"success": False, "message": "Database connection error. Please try again."}), 500
        validated_deleted_ids = set()

        for raw_deleted_id in deleted_ids:
            try:
                deleted_id = int(raw_deleted_id)
            except (TypeError, ValueError):
                return jsonify({"success": False, "message": f"Invalid deleted row id: {raw_deleted_id}"}), 400

            # Ignore client-side temporary ids (non-positive values)
            if deleted_id <= 0:
                continue

            if deleted_id not in existing:
                return jsonify({"success": False, "message": f"Deleted row id {deleted_id} not found."}), 400

            validated_deleted_ids.add(deleted_id)

        seen_numbers = set()
        validated_rows = []
        errors = []

        for idx, row in enumerate(rows):
            row_id = row.get("id")
            try:
                consignment_number = normalize_consignment_number(row.get("consignment_number"))
            except ValueError as error:
                errors.append({"index": idx, "field": "consignment_number", "message": str(error)})
                consignment_number = None

            try:
                status = normalize_status(row.get("status"))
            except ValueError as error:
                errors.append({"index": idx, "field": "status", "message": str(error)})
                status = None

            try:
                pickup_pincode = normalize_indian_pincode(row.get("pickup_pincode"), "pickup_pincode")
            except ValueError as error:
                errors.append({"index": idx, "field": "pickup_pincode", "message": str(error)})
                pickup_pincode = None

            try:
                drop_pincode = normalize_indian_pincode(row.get("drop_pincode"), "drop_pincode")
            except ValueError as error:
                errors.append({"index": idx, "field": "drop_pincode", "message": str(error)})
                drop_pincode = None

            eta = str(row.get("eta") or "").strip()

            if consignment_number:
                if consignment_number in seen_numbers:
                    errors.append({"index": idx, "field": "consignment_number", "message": f"Duplicate consignment number in sheet: {consignment_number}"})
                seen_numbers.add(consignment_number)

            if row_id is not None:
                try:
                    row_id = int(row_id)
                except (TypeError, ValueError):
                    errors.append({"index": idx, "field": "id", "message": f"Invalid row id: {row_id}"})
                    row_id = None

                # Treat non-positive ids (client temporary ids like -1) as new rows
                if row_id and row_id <= 0:
                    row_id = None
                else:
                    if row_id and row_id not in existing:
                        errors.append({"index": idx, "field": "id", "message": f"Row id {row_id} not found."})
                        row_id = None

                    if row_id and row_id in validated_deleted_ids:
                        errors.append({"index": idx, "field": "id", "message": f"Row id {row_id} cannot be updated and deleted in the same save."})
            else:
                row_id = None

            pickup_tag = str(row.get("pickup_tag") or "").strip()
            pickup_date = str(row.get("pickup_date") or "").strip()
            drop_tag = str(row.get("drop_tag") or "").strip()
            drop_date = str(row.get("drop_date") or "").strip()
            pod_file_data = str(row.get("pod_file_data") or "").strip() or None
            pod_file_name = str(row.get("pod_file_name") or "").strip() or None
            pod_file_type = str(row.get("pod_file_type") or "").strip() or None
            pod_image = str(row.get("pod_image") or "").strip() or None

            # Never persist external POD URLs. They are often presigned and expire,
            # which breaks long-term archive requirements.
            if pod_image and _is_external_pod_url(pod_image):
                errors.append({
                    "index": idx,
                    "field": "pod_image",
                    "message": "External POD URLs are not allowed. Upload the image file so we can store a permanent reference.",
                })

            validated_rows.append({
                "id": row_id,
                "consignment_number": consignment_number,
                "status": status,
                "pickup_pincode": pickup_pincode,
                "pickup_address": str(row.get("pickup_address") or "").strip(),
                "pickup_tag": pickup_tag,
                "pickup_date": pickup_date,
                "drop_pincode": drop_pincode,
                "drop_address": str(row.get("drop_address") or "").strip(),
                "drop_tag": drop_tag,
                "drop_date": drop_date,
                "eta": eta,
                "pod_image": pod_image,
                "pod_file_data": pod_file_data,
                "pod_file_name": pod_file_name,
                "pod_file_type": pod_file_type,
            })

        # If any per-row validation errors were collected, return them instead of aborting.
        if errors:
            db.session.rollback()
            return jsonify({"success": False, "errors": errors}), 400

        # Detect missing DB columns proactively to avoid failing mid-commit with unclear errors.
        try:
            inspector = sa_inspect(db.engine)
            db_columns = {c['name'] for c in inspector.get_columns('consignment')}
            model_columns = {col.name for col in Consignment.__table__.columns}
            missing_cols = model_columns.difference(db_columns)
            if missing_cols:
                logger.exception("Schema mismatch detected while saving: missing columns %s", missing_cols)
                db.session.rollback()
                return jsonify({"success": False, "message": "Database schema needs an update. Missing consignment fields."}), 500
        except Exception:
            # If inspection fails for any reason, continue and let DB operations surface errors.
            logger.exception("Failed to inspect consignment table columns")

        for deleted_id in validated_deleted_ids:
            consignment = db.session.get(Consignment, int(deleted_id))
            if consignment is not None:
                db.session.delete(consignment)

        for row in validated_rows:
            if row["id"]:
                # Load the specific Consignment instance on demand. This avoids a full-table
                # select which may fail when the DB schema differs from model mappings.
                try:
                    consignment = db.session.get(Consignment, int(row["id"]))
                except Exception:
                    consignment = None

                if consignment is None:
                    # This should generally not happen (validated earlier), but handle gracefully.
                    consignment = Consignment()
                    db.session.add(consignment)
            else:
                consignment = Consignment()
                db.session.add(consignment)

            previous_pod_image = getattr(consignment, 'pod_image', None)
            new_pod_image = row.get("pod_image")

            if row.get("pod_file_data"):
                try:
                    pod_bytes = _decode_pod_data_url(row["pod_file_data"])
                    original_name = row.get("pod_file_name") or "pod.jpg"
                    filename = f"{uuid.uuid4().hex}_{secure_filename(original_name)}"
                    new_pod_image = _store_pod_bytes(filename, pod_bytes, row.get("pod_file_type"))
                except ValueError as error:
                    db.session.rollback()
                    return jsonify({"success": False, "message": str(error)}), 400

            if previous_pod_image and new_pod_image and previous_pod_image != new_pod_image:
                _delete_pod_file(previous_pod_image)

            consignment.consignment_number = row["consignment_number"]
            consignment.status = row["status"]
            consignment.pickup_pincode = row["pickup_pincode"]
            consignment.pickup_address = row.get("pickup_address")
            consignment.pickup_tag = row.get("pickup_tag")
            consignment.pickup_date = row.get("pickup_date")
            consignment.drop_pincode = row["drop_pincode"]
            consignment.drop_address = row.get("drop_address")
            consignment.drop_tag = row.get("drop_tag")
            consignment.drop_date = row.get("drop_date")
            consignment.eta = row["eta"]
            consignment.pod_image = new_pod_image

        db.session.commit()
        # Return the updated total so the client can navigate to the page
        # that will contain newly inserted rows (the new last page).
        try:
            total = Consignment.query.count()
        except Exception:
            total = None

        return jsonify({
            "success": True,
            "message": "Sheet saved successfully.",
            "deleted_count": len(validated_deleted_ids),
            "total": total,
        })

    except IntegrityError:
        db.session.rollback()
        logger.exception("Integrity error in admin save")
        return jsonify({"success": False, "message": "Duplicate consignment number already exists."}), 400
    except Exception:
        db.session.rollback()
        logger.exception("Unexpected error in admin save")
        return jsonify({"success": False, "message": "An unexpected error occurred. Please try again."}), 500


@admin_bp.route("/admin/consignments/archive", methods=["POST"], endpoint="consignments_archive")
@limiter.limit("10 per minute")
@require_admin
def consignments_archive():
    payload = request.get_json(silent=True) or {}
    before_date = str(payload.get("before_date") or "").strip()

    if not before_date:
        return jsonify({"success": False, "message": "Please provide a cutoff date."}), 400

    cutoff_date = _parse_date_string(before_date)
    if cutoff_date is None:
        return jsonify({"success": False, "message": "Cutoff date must be a valid date in YYYY-MM-DD format."}), 400

    try:
        query = Consignment.query.filter(Consignment.status.ilike("Delivered"))
        query = query.filter(Consignment.drop_date.isnot(None), Consignment.drop_date != "")

        archived_count = 0
        for consignment in query.all():
            drop_date = _parse_date_string(getattr(consignment, "drop_date", ""))
            if drop_date is None:
                continue
            if drop_date < cutoff_date:
                if getattr(consignment, "pod_image", None):
                    _delete_pod_file(consignment.pod_image)
                db.session.delete(consignment)
                archived_count += 1

        db.session.commit()
        return jsonify({"success": True, "archived_count": archived_count})
    except (OperationalError, DatabaseError):
        db.session.rollback()
        logger.exception("Database error archiving consignments")
        return jsonify({"success": False, "message": "Unable to archive consignments. Please try again."}), 500
    except Exception:
        db.session.rollback()
        logger.exception("Unexpected error archiving consignments")
        return jsonify({"success": False, "message": "An unexpected error occurred while archiving consignments."}), 500


@admin_bp.route("/admin/consignments/import", methods=["POST"], endpoint="consignments_import_excel")
@limiter.limit("10 per minute")
@require_admin
def consignments_import_excel():
    upload = request.files.get("file")
    if not upload or not upload.filename:
        flash("Please choose an Excel file (.xlsx).", "danger")
        return redirect(url_for("admin.consignments_panel"))

    filename = upload.filename.lower()
    if not filename.endswith(".xlsx"):
        flash("Only .xlsx files are supported.", "danger")
        return redirect(url_for("admin.consignments_panel"))

    try:
        workbook = load_workbook(upload, data_only=True)
        sheet = workbook.active

        header_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_cells:
            flash("Excel file is empty.", "danger")
            return redirect(url_for("admin.consignments_panel"))

        normalized_headers = [_normalize_header(cell) for cell in header_cells]
        header_index = {name: idx for idx, name in enumerate(normalized_headers) if name}

        consignment_idx = header_index.get("consignment_number")
        status_idx = header_index.get("status")
        pickup_address_idx = header_index.get("pickup_address")
        pickup_pincode_idx = header_index.get("pickup_pincode")
        pickup_tag_idx = header_index.get("pickup_tag")
        pickup_date_idx = header_index.get("pickup_date")
        drop_address_idx = header_index.get("drop_address")
        drop_pincode_idx = header_index.get("drop_pincode")
        drop_tag_idx = header_index.get("drop_tag")
        drop_date_idx = header_index.get("drop_date")
        eta_idx = header_index.get("eta")

        # Required headers: consignment_number and status.
        if None in (consignment_idx, status_idx):
            flash("Required headers: consignment_number, status", "danger")
            return redirect(url_for("admin.consignments_panel"))

        existing_numbers = {c.consignment_number for c in Consignment.query.with_entities(Consignment.consignment_number).all()}
        file_seen = set()
        added_count = 0
        skipped_count = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or all(value is None or str(value).strip() == "" for value in row):
                continue

            consignment_number = normalize_consignment_number(row[consignment_idx])
            status = normalize_status(row[status_idx])
            pickup_address = str(row[pickup_address_idx] if pickup_address_idx is not None and row[pickup_address_idx] is not None else "").strip()
            pickup_pincode = normalize_indian_pincode(row[pickup_pincode_idx] if pickup_pincode_idx is not None and row[pickup_pincode_idx] is not None else "", "pickup_pincode")
            pickup_tag = str(row[pickup_tag_idx] if pickup_tag_idx is not None and row[pickup_tag_idx] is not None else "").strip()
            pickup_date = str(row[pickup_date_idx] if pickup_date_idx is not None and row[pickup_date_idx] is not None else "").strip()
            drop_address = str(row[drop_address_idx] if drop_address_idx is not None and row[drop_address_idx] is not None else "").strip()
            drop_pincode = normalize_indian_pincode(row[drop_pincode_idx] if drop_pincode_idx is not None and row[drop_pincode_idx] is not None else "", "drop_pincode")
            drop_tag = str(row[drop_tag_idx] if drop_tag_idx is not None and row[drop_tag_idx] is not None else "").strip()
            drop_date = str(row[drop_date_idx] if drop_date_idx is not None and row[drop_date_idx] is not None else "").strip()
            eta = str(row[eta_idx] if eta_idx is not None and row[eta_idx] is not None else "").strip()

            if consignment_number in existing_numbers or consignment_number in file_seen:
                skipped_count += 1
                continue

            consignment = Consignment(
                consignment_number=consignment_number,
                status=status,
                pickup_address=pickup_address,
                pickup_pincode=pickup_pincode,
                pickup_tag=pickup_tag,
                pickup_date=pickup_date,
                drop_address=drop_address,
                drop_pincode=drop_pincode,
                drop_tag=drop_tag,
                drop_date=drop_date,
                eta=eta,
            )

            db.session.add(consignment)
            file_seen.add(consignment_number)
            existing_numbers.add(consignment_number)
            added_count += 1

        db.session.commit()
        flash(f"Import completed. Added: {added_count}, skipped duplicates: {skipped_count}.", "success")
        return redirect(url_for("admin.consignments_panel"))
    except ValueError as error:
        db.session.rollback()
        flash(str(error), "danger")
        return redirect(url_for("admin.consignments_panel"))
    except Exception:
        db.session.rollback()
        logger.exception("Unexpected error in Excel import")
        flash("Failed to import Excel file.", "danger")
        return redirect(url_for("admin.consignments_panel"))


@admin_bp.route("/admin/consignments/export.xlsx", methods=["GET"], endpoint="consignments_export_excel")
@require_admin
def consignments_export_excel():
    try:
        rows = Consignment.query.order_by(Consignment.id.asc()).all()

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Internal Consignments"

        sheet.append([
            "consignment_number",
            "status",
            "pickup_tag",
            "drop_pincode",
            "pickup_date",
            "drop_date",
            "pickup_address",
            "drop_address",
        ])

        for row in rows:
            sheet.append([
                row.consignment_number,
                row.status,
                getattr(row, "pickup_tag", ""),
                row.drop_pincode,
                getattr(row, "pickup_date", ""),
                getattr(row, "drop_date", ""),
                getattr(row, "pickup_address", ""),
                getattr(row, "drop_address", ""),
            ])

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="internal_consignments.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception:
        logger.exception("Excel export failed")
        return jsonify({"success": False, "message": "Failed to export Excel."}), 500


@admin_bp.route("/admin/consignments/export.pdf", methods=["GET"], endpoint="consignments_export_pdf")
@require_admin
def consignments_export_pdf():
    try:
        rows = Consignment.query.order_by(Consignment.id.asc()).all()

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
        styles = getSampleStyleSheet()

        table_data = [["Consignment #", "Status", "Pickup Tag", "Drop Pin", "Pickup Date", "Drop Estimated"]]
        for row in rows:
            table_data.append([
                row.consignment_number or "",
                row.status or "",
                getattr(row, "pickup_tag", "") or "",
                row.drop_pincode or "",
                getattr(row, "pickup_date", "") or "",
                getattr(row, "drop_date", "") or "",
            ])

        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E9ECEF")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#6C757D")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))

        content = [
            Paragraph("Internal Consignment MIS", styles["Heading2"]),
            Spacer(1, 8),
            table,
        ]

        doc.build(content)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="internal_consignments.pdf",
            mimetype="application/pdf",
        )
    except Exception:
        logger.exception("PDF export failed")
        return jsonify({"success": False, "message": "Failed to export PDF."}), 500


@admin_bp.route("/admin/consignments/import-template.xlsx", methods=["GET"], endpoint="consignments_import_template_excel")
@require_admin
def consignments_import_template_excel():
    try:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Import Template"

        sheet.append([
            "consignment_number",
            "status",
            "pickup_address",
            "pickup_pincode",
            "pickup_tag",
            "pickup_date",
            "drop_address",
            "drop_pincode",
            "drop_tag",
            "drop_date",
        ])

        sheet.append([
            "CN001",
            "In Transit",
            "123 Main Street, New Delhi",
            "110017",
            "PICKUP-001",
            "2026-05-10",
            "456 Marine Drive, Mumbai",
            "400001",
            "DROP-001",
            "2026-05-12",
        ])

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="internal_consignments_import_template.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception:
        logger.exception("Template export failed")
        return jsonify({"success": False, "message": "Failed to generate import template."}), 500


@admin_bp.route("/admin/consignments/<int:consignment_id>/pod", methods=["GET"], endpoint="consignment_pod_file")
@require_admin
def consignment_pod_file(consignment_id):
    """Serve the POD file for a consignment if present."""
    try:
        consignment = db.session.get(Consignment, consignment_id)
        if not consignment or not getattr(consignment, "pod_image", None):
            return jsonify({"success": False, "message": "No POD found."}), 404

        # pod_image may be:
        # - a local relative filename stored under instance/uploads
        # - a supabase marker value like "supabase:bucket/path/to/file"
        # - a public URL (http/https)
        pod_path = consignment.pod_image
        if pod_path.startswith("http://") or pod_path.startswith("https://"):
            return redirect(pod_path)

        # Supabase-stored value: "supabase:bucket/path". Keep the app route as the
        # stable POD URL and stream bytes directly instead of redirecting to a
        # temporary storage URL.
        if isinstance(pod_path, str) and pod_path.startswith("supabase:"):
            try:
                content, object_path = _download_supabase_pod_file(pod_path)
                mimetype, _ = mimetypes.guess_type(object_path)
                return send_file(io.BytesIO(content), mimetype=mimetype or "application/octet-stream")
            except RuntimeError as error:
                logger.exception("Error downloading Supabase POD file")
                return jsonify({"success": False, "message": str(error)}), 500
            except Exception:
                logger.exception("Error serving Supabase POD file")
                return jsonify({"success": False, "message": "Failed to serve POD."}), 500

        # Otherwise treat as local filename under instance/uploads
        upload_folder = os.path.join(current_app.instance_path, "uploads")
        safe_path = os.path.normpath(os.path.join(upload_folder, pod_path))
        if not safe_path.startswith(os.path.abspath(upload_folder)):
            return jsonify({"success": False, "message": "Invalid POD path."}), 400

        if not os.path.exists(safe_path):
            try:
                content, bucket, object_path = _download_legacy_supabase_pod_file(consignment.id, pod_path)
                consignment.pod_image = f"supabase:{bucket}/{object_path}"
                db.session.commit()
                mimetype, _ = mimetypes.guess_type(object_path)
                return send_file(io.BytesIO(content), mimetype=mimetype or "application/octet-stream")
            except Exception:
                db.session.rollback()
                logger.exception("Legacy local POD was not found locally or in Supabase")
                return jsonify({"success": False, "message": "POD file missing."}), 404

        return send_file(safe_path)
    except Exception:
        logger.exception("Error serving POD file")
        return jsonify({"success": False, "message": "Failed to serve POD."}), 500


@admin_bp.route("/admin/consignments/<int:consignment_id>/pod", methods=["POST"], endpoint="consignment_pod_upload")
@require_admin
def consignment_pod_upload(consignment_id):
    """Upload or replace a POD image for a consignment. Returns JSON with new pod path/url."""
    upload = request.files.get("file")
    if not upload or not upload.filename:
        return jsonify({"success": False, "message": "No file uploaded."}), 400

    if not (upload.mimetype or "").startswith("image/"):
        return jsonify({"success": False, "message": "POD must be an image file."}), 400

    try:
        consignment = db.session.get(Consignment, consignment_id)
        if not consignment:
            return jsonify({"success": False, "message": "Consignment not found."}), 404

        filename = secure_filename(upload.filename)
        filename = f"{uuid.uuid4().hex}_{filename}"
        file_bytes = upload.read()
        if len(file_bytes) > MAX_POD_IMAGE_BYTES:
            return jsonify({"success": False, "message": "POD image must be smaller than 5 MB."}), 400

        # If Supabase configured, upload there and store a marker 'supabase:bucket/path'
        supa = _get_supabase_client()
        bucket = os.getenv('SUPABASE_BUCKET', 'pod-uploads')
        if supa:
            try:
                object_path = f"{consignment_id}/{filename}"
                # upload to supabase; storage client expects a file path, so write temp file
                tmp = None
                try:
                    tf = tempfile.NamedTemporaryFile(delete=False)
                    tmp = tf.name
                    tf.write(file_bytes)
                    tf.flush()
                    tf.close()
                    supa.storage.from_(bucket).upload(object_path, tmp, {'content-type': upload.mimetype or 'application/octet-stream'})
                finally:
                    try:
                        if tmp and os.path.exists(tmp):
                            os.remove(tmp)
                    except Exception:
                        logger.exception('Failed to remove temporary POD upload file')

                # store marker so we can remove later
                consignment.pod_image = f"supabase:{bucket}/{object_path}"
                db.session.commit()
                return jsonify({"success": True, "pod_image": consignment.pod_image}), 200
            except Exception:
                db.session.rollback()
                logger.exception("Supabase POD upload failed; falling back to local storage")

        # fallback: local instance storage
        try:
            upload_folder = os.path.join(current_app.instance_path, "uploads")
            os.makedirs(upload_folder, exist_ok=True)
            dest_path = os.path.join(upload_folder, filename)
            with open(dest_path, "wb") as file_handle:
                file_handle.write(file_bytes)
            consignment.pod_image = filename
            db.session.commit()
            return jsonify({"success": True, "pod_image": filename}), 200
        except Exception:
            db.session.rollback()
            logger.exception("POD upload failed (local)")
            return jsonify({"success": False, "message": "Upload failed."}), 500
    except Exception:
        db.session.rollback()
        logger.exception("POD upload failed")
        return jsonify({"success": False, "message": "Upload failed."}), 500


@admin_bp.route("/admin/consignments/<int:consignment_id>/pod", methods=["DELETE"], endpoint="consignment_pod_delete")
@require_admin
def consignment_pod_delete(consignment_id):
    """Delete POD association (and file if present)."""
    try:
        consignment = db.session.get(Consignment, consignment_id)
        if not consignment or not getattr(consignment, "pod_image", None):
            return jsonify({"success": False, "message": "No POD to delete."}), 404

        # If stored in Supabase, remove via storage API
        pod_val = consignment.pod_image
        if isinstance(pod_val, str) and pod_val.startswith('supabase:'):
            client = _get_supabase_client()
            if client:
                try:
                    _, rest = pod_val.split(":", 1)
                    bucket, object_path = rest.split("/", 1)
                    client.storage.from_(bucket).remove([object_path])
                except Exception:
                    logger.exception("Failed to remove POD from Supabase")

        else:
            # local file
            upload_folder = os.path.join(current_app.instance_path, "uploads")
            pod_rel = consignment.pod_image
            try:
                pod_path = os.path.normpath(os.path.join(upload_folder, pod_rel))
                if pod_path.startswith(os.path.abspath(upload_folder)) and os.path.exists(pod_path):
                    try:
                        os.remove(pod_path)
                    except Exception:
                        logger.exception("Failed to remove POD file from disk")
            except Exception:
                logger.exception("Error while attempting to remove local POD file")

        consignment.pod_image = None
        db.session.commit()
        return jsonify({"success": True}), 200
    except Exception:
        db.session.rollback()
        logger.exception("Failed deleting POD")
        return jsonify({"success": False, "message": "Delete failed."}), 500
