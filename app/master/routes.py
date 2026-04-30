import logging
import re
from datetime import datetime
from io import BytesIO

from flask import Blueprint, flash, jsonify, redirect, render_template, request, url_for
from openpyxl import load_workbook
from sqlalchemy import asc, desc, select
from sqlalchemy.dialects.postgresql import insert as pg_insert

from app import limiter
from app.admin.auth import require_admin
from app.eta_master.models import EtaMasterRecord, PickupStation
from app.models import Consignment, db

logger = logging.getLogger(__name__)

master_bp = Blueprint('master', __name__, template_folder='templates', url_prefix='/master')

# ============================================================================
# PICKUP STATIONS
# ============================================================================


@master_bp.route('/pickup-stations', methods=['GET'], endpoint='pickup_stations_panel')
@require_admin
def pickup_stations_panel():
    stations = PickupStation.query.order_by(PickupStation.name.asc()).all()
    return render_template('master/pickup_stations.html', pickup_stations=[s.to_dict() for s in stations], active_page='eta_master')


@master_bp.route('/pickup-stations/list', methods=['GET'], endpoint='pickup_stations_list')
def pickup_stations_list():
    stations = PickupStation.query.order_by(PickupStation.name.asc()).all()
    return jsonify([s.name for s in stations])


@master_bp.route('/pickup-stations', methods=['POST'], endpoint='pickup_stations_create')
@require_admin
def pickup_stations_create():
    payload = request.get_json(silent=True) or {}
    name = (payload.get('name') or '').strip()
    pin = (payload.get('pin_code') or '').strip()
    address = (payload.get('address') or '').strip() or None

    if not name or not pin:
        return jsonify({'success': False, 'message': 'Name and pin_code are required.'}), 400

    if not re.fullmatch(r'[1-9][0-9]{5}', pin):
        return jsonify({'success': False, 'message': 'pin_code must be a valid 6-digit pincode.'}), 400

    existing = PickupStation.query.filter_by(name=name).first()
    if existing:
        return jsonify({'success': False, 'message': 'Pickup station with that name already exists.'}), 400

    station = PickupStation(name=name, pin_code=pin, address=address)
    db.session.add(station)
    db.session.commit()
    return jsonify({'success': True, 'station': station.to_dict()})


@master_bp.route('/pickup-stations/<int:station_id>', methods=['PUT'], endpoint='pickup_stations_update')
@require_admin
def pickup_stations_update(station_id):
    payload = request.get_json(silent=True) or {}
    station = db.session.get(PickupStation, station_id)
    if not station:
        return jsonify({'success': False, 'message': 'Station not found.'}), 404

    name = payload.get('name')
    pin = payload.get('pin_code')
    address = payload.get('address')

    if name:
        station.name = str(name).strip()
    if pin:
        pin = str(pin).strip()
        if not re.fullmatch(r'[1-9][0-9]{5}', pin):
            return jsonify({'success': False, 'message': 'pin_code must be a valid 6-digit pincode.'}), 400
        station.pin_code = pin
    station.address = str(address).strip() or None

    db.session.commit()
    return jsonify({'success': True, 'station': station.to_dict()})


@master_bp.route('/pickup-stations/<int:station_id>', methods=['DELETE'], endpoint='pickup_stations_delete')
@require_admin
def pickup_stations_delete(station_id):
    station = db.session.get(PickupStation, station_id)
    if not station:
        return jsonify({'success': False, 'message': 'Station not found.'}), 404
    db.session.delete(station)
    db.session.commit()
    return jsonify({'success': True})


# ============================================================================
# ETA MASTER (moved from eta_master blueprint)
# ============================================================================


HEADER_ALIASES = {
    # SNO variations
    'sno': 'sno',
    'serial number': 'sno',
    'serial no': 'sno',
    's.no': 'sno',
    'slno': 'sno',
    
    # PIN CODE variations
    'pin code': 'pin_code',
    'pincode': 'pin_code',
    'pin': 'pin_code',
    'postal code': 'pin_code',
    'zip code': 'pin_code',
    
    # PICKUP STATION variations
    'pickup station': 'pickup_station',
    'pick up station': 'pickup_station',
    'pickup stn': 'pickup_station',
    'pick-up station': 'pickup_station',
    'pickupstation': 'pickup_station',
    'station': 'pickup_station',
    
    # STATE/UT variations
    'state/ut': 'state_ut',
    'state ut': 'state_ut',
    'state': 'state_ut',
    'state/union territory': 'state_ut',
    'statut': 'state_ut',
    'state-ut': 'state_ut',
    
    # CITY variations
    'city': 'city',
    'city name': 'city',
    'destination city': 'city',
    
    # PICKUP LOCATION variations
    'pickup location': 'pickup_location',
    'pick up location': 'pickup_location',
    'pick-up location': 'pickup_location',
    'pickup loc': 'pickup_location',
    'pickuplocation': 'pickup_location',
    'pick up': 'pickup_location',
    'pickup': 'pickup_location',
    'source location': 'pickup_location',
    
    # DELIVERY LOCATION variations
    'delivery location': 'delivery_location',
    'delivery loc': 'delivery_location',
    'deliverylocation': 'delivery_location',
    'delivery': 'delivery_location',
    'destination location': 'delivery_location',
    'drop location': 'delivery_location',
    
    # TAT IN DAYS variations
    'tat in days': 'tat_in_days',
    'tat': 'tat_in_days',
    'tat (days)': 'tat_in_days',
    'tat days': 'tat_in_days',
    'delivery time (days)': 'tat_in_days',
    'time to deliver': 'tat_in_days',
    'turnaround time': 'tat_in_days',
    
    # ZONE variations
    'zone': 'zone',
    'region': 'zone',
    'area': 'zone',
    'zone code': 'zone',
}

REQUIRED_FIELDS = ['pin_code', 'pickup_station', 'state_ut', 'city', 'pickup_location', 'delivery_location', 'tat_in_days', 'zone']


def _normalize_header(value):
    return re.sub(r'\s+', ' ', str(value or '').strip().lower())


def _map_headers(header_row):
    mapped = {}
    for index, header in enumerate(header_row):
        normalized = _normalize_header(header)
        mapped[index] = HEADER_ALIASES.get(normalized)
    return mapped


def _cast_sno(value):
    """Cast SNO to Integer or None."""
    if value is None or value == '' or str(value).strip() == '':
        return None
    try:
        return int(float(value))  # float first to handle Excel decimals
    except (TypeError, ValueError):
        raise ValueError(f'SNO must be an integer, got: {value}')


def _cast_pincode(value):
    """Cast PIN CODE to String (6 digits exactly)."""
    text = str(value or '').strip()
    if not text:
        raise ValueError('PIN CODE is required.')
    if not re.fullmatch(r'\d{6}', text):
        raise ValueError(f'PIN CODE must be exactly 6 digits, got: {text}')
    return text


def _cast_pickup_station(value):
    """Cast PICK UP STATION to String (required, max 255 chars)."""
    text = str(value or '').strip()
    if not text:
        raise ValueError('PICK UP STATION is required.')
    if len(text) > 255:
        raise ValueError(f'PICK UP STATION exceeds 255 characters: {len(text)} chars')
    return text


def _cast_state_ut(value):
    """Cast STATE/UT to String (required, max 100 chars)."""
    text = str(value or '').strip()
    if not text:
        raise ValueError('STATE/UT is required.')
    if len(text) > 100:
        raise ValueError(f'STATE/UT exceeds 100 characters: {len(text)} chars')
    return text


def _cast_city(value):
    """Cast CITY to String (required, max 100 chars)."""
    text = str(value or '').strip()
    if not text:
        raise ValueError('CITY is required.')
    if len(text) > 100:
        raise ValueError(f'CITY exceeds 100 characters: {len(text)} chars')
    return text


def _cast_pickup_location(value):
    """Cast PICK UP LOCATION to String (required, max 255 chars)."""
    text = str(value or '').strip()
    if not text:
        raise ValueError('PICK UP LOCATION is required.')
    if len(text) > 255:
        raise ValueError(f'PICK UP LOCATION exceeds 255 characters: {len(text)} chars')
    return text


def _cast_delivery_location(value):
    """Cast DELIVERY LOCATION to String (required, max 255 chars)."""
    text = str(value or '').strip()
    if not text:
        raise ValueError('DELIVERY LOCATION is required.')
    if len(text) > 255:
        raise ValueError(f'DELIVERY LOCATION exceeds 255 characters: {len(text)} chars')
    return text


def _cast_tat_in_days(value):
    """Cast TAT IN DAYS to Float (required, non-negative)."""
    if value is None or value == '' or str(value).strip() == '':
        raise ValueError('TAT IN DAYS is required.')
    try:
        tat = float(value)
    except (TypeError, ValueError):
        raise ValueError(f'TAT IN DAYS must be numeric, got: {value}')
    if tat < 0:
        raise ValueError(f'TAT IN DAYS cannot be negative, got: {tat}')
    return tat


def _cast_zone(value):
    """Cast ZONE to String (required, max 50 chars)."""
    text = str(value or '').strip()
    if not text:
        raise ValueError('ZONE is required.')
    if len(text) > 50:
        raise ValueError(f'ZONE exceeds 50 characters: {len(text)} chars')
    return text


def _build_record_payload(source):
    """Validate and normalize a record payload from a form or dict-like source."""
    sno = _cast_sno(source.get('sno'))
    pin_code = _cast_pincode(source.get('pin_code'))
    pickup_station = _cast_pickup_station(source.get('pickup_station'))
    state_ut = _cast_state_ut(source.get('state_ut'))
    city = _cast_city(source.get('city'))
    pickup_location = _cast_pickup_location(source.get('pickup_location'))
    delivery_location = _cast_delivery_location(source.get('delivery_location'))
    tat_in_days = _cast_tat_in_days(source.get('tat_in_days'))
    zone = _cast_zone(source.get('zone'))

    record_key = EtaMasterRecord.build_record_key(
        pin_code,
        pickup_station,
        state_ut,
        city,
        pickup_location,
        delivery_location,
        zone,
    )

    now = datetime.utcnow()
    return {
        'record_key': record_key,
        'sno': sno,
        'pin_code': pin_code,
        'pickup_station': pickup_station,
        'state_ut': state_ut,
        'city': city,
        'pickup_location': pickup_location,
        'delivery_location': delivery_location,
        'tat_in_days': tat_in_days,
        'zone': zone,
        'source_filename': source.get('source_filename'),
        'source_row_number': source.get('source_row_number'),
        'imported_at': now,
        'updated_at': now,
    }


def _upsert_records(rows, source_filename):
    """Bulk upsert ETA master rows with PostgreSQL ON CONFLICT handling."""
    inserted = 0
    updated = 0
    skipped = 0
    errors = []

    records = []
    for row_number, row in rows:
        try:
            record_data = _build_record_payload(row)
        except ValueError as error:
            errors.append({'row': row_number, 'error': str(error)})
            continue
        except Exception as error:
            errors.append({'row': row_number, 'error': f'Unexpected row error: {error}'})
            continue

        record_data['source_filename'] = source_filename
        record_data['source_row_number'] = row_number
        records.append(record_data)

    if not records:
        return inserted, updated, skipped, errors

    try:
        record_keys = [record['record_key'] for record in records]
        existing_keys = {
            key for (key,) in db.session.query(EtaMasterRecord.record_key)
            .filter(EtaMasterRecord.record_key.in_(record_keys))
            .all()
        }
        inserted = sum(1 for record in records if record['record_key'] not in existing_keys)
        updated = len(records) - inserted

        insert_stmt = pg_insert(EtaMasterRecord).values(records)
        upsert_stmt = insert_stmt.on_conflict_do_update(
            index_elements=['record_key'],
            set_={
                'sno': insert_stmt.excluded.sno,
                'pin_code': insert_stmt.excluded.pin_code,
                'pickup_station': insert_stmt.excluded.pickup_station,
                'state_ut': insert_stmt.excluded.state_ut,
                'city': insert_stmt.excluded.city,
                'pickup_location': insert_stmt.excluded.pickup_location,
                'delivery_location': insert_stmt.excluded.delivery_location,
                'tat_in_days': insert_stmt.excluded.tat_in_days,
                'zone': insert_stmt.excluded.zone,
                'source_filename': insert_stmt.excluded.source_filename,
                'source_row_number': insert_stmt.excluded.source_row_number,
                'updated_at': datetime.utcnow(),
            },
        )
        db.session.execute(upsert_stmt)
        db.session.commit()
        logger.info("Bulk upsert complete: %s inserted, %s updated", inserted, updated)
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error during bulk upsert: {e}")
        errors.append({'row': 'batch', 'error': f'Database error: {str(e)}'})

    return inserted, updated, skipped, errors


SORTABLE_COLUMNS = {
    'sno': EtaMasterRecord.sno,
    'pin_code': EtaMasterRecord.pin_code,
    'pickup_station': EtaMasterRecord.pickup_station,
    'city': EtaMasterRecord.city,
    'state_ut': EtaMasterRecord.state_ut,
    'pickup_location': EtaMasterRecord.pickup_location,
    'delivery_location': EtaMasterRecord.delivery_location,
    'tat_in_days': EtaMasterRecord.tat_in_days,
    'zone': EtaMasterRecord.zone,
}


def _get_pagination_params():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 100, type=int)
    return max(page, 1), max(1, min(per_page, 500))


def _get_search_params():
    search = request.args.get('search', '').strip()
    search_type = request.args.get('search_type', 'pin_code').strip().lower()
    if search_type not in ('pin_code', 'consignment', 'sno'):
        search_type = 'pin_code'
    sort_by = request.args.get('sort_by', '').strip().lower()
    if sort_by not in SORTABLE_COLUMNS:
        sort_by = ''
    sort_dir = request.args.get('sort_dir', 'asc').strip().lower()
    if sort_dir not in ('asc', 'desc'):
        sort_dir = 'asc'
    return search, search_type, sort_by, sort_dir


def _resolve_search_pincodes(search, search_type):
    """Return a list of pin codes to filter on, or None if not applicable."""
    if not search:
        return None
    if search_type == 'pin_code':
        return [search]
    if search_type == 'consignment':
        consignment = Consignment.query.filter_by(consignment_number=search).first()
        if not consignment:
            return []  # empty list → no results
        pincodes = []
        if consignment.pickup_pincode:
            pincodes.append(consignment.pickup_pincode)
        if consignment.drop_pincode:
            pincodes.append(consignment.drop_pincode)
        return pincodes if pincodes else []
    if search_type == 'sno':
        return None  # handled separately
    return None


def _build_eta_query(search, search_type, sort_by, sort_dir):
    """Build the SQLAlchemy SELECT for ETA master with optional search/sort."""
    stmt = select(EtaMasterRecord)

    if search:
        if search_type == 'sno':
            try:
                sno_val = int(search)
                stmt = stmt.where(EtaMasterRecord.sno == sno_val)
            except ValueError:
                stmt = stmt.where(EtaMasterRecord.id == -1)  # non-numeric SNO → empty result
        else:
            pincodes = _resolve_search_pincodes(search, search_type)
            if pincodes is None:
                pass  # no filter
            elif len(pincodes) == 0:
                stmt = stmt.where(EtaMasterRecord.id == -1)  # force empty result
            elif len(pincodes) == 1:
                stmt = stmt.where(EtaMasterRecord.pin_code == pincodes[0])
            else:
                stmt = stmt.where(EtaMasterRecord.pin_code.in_(pincodes))

    if sort_by and sort_by in SORTABLE_COLUMNS:
        col = SORTABLE_COLUMNS[sort_by]
        stmt = stmt.order_by(asc(col) if sort_dir == 'asc' else desc(col))
    else:
        stmt = stmt.order_by(EtaMasterRecord.id.desc())

    return stmt


def _paginate_eta_master(page, per_page, search='', search_type='pin_code', sort_by='', sort_dir='asc'):
    stmt = _build_eta_query(search, search_type, sort_by, sort_dir)
    pagination = db.paginate(stmt, page=page, per_page=per_page, error_out=False)

    if pagination.total and page > pagination.pages:
        pagination = db.paginate(stmt, page=pagination.pages, per_page=per_page, error_out=False)
    return pagination


@master_bp.route('/eta', methods=['GET'], endpoint='eta_master_upload')
@require_admin
def eta_master_upload():
    page, per_page = _get_pagination_params()
    search, search_type, sort_by, sort_dir = _get_search_params()
    pagination = _paginate_eta_master(page, per_page, search, search_type, sort_by, sort_dir)
    return render_template(
        'master/eta_master.html',
        records=pagination.items,
        total=pagination.total,
        pages=pagination.pages,
        current_page=page,
        per_page=per_page,
        search=search,
        search_type=search_type,
        sort_by=sort_by,
        sort_dir=sort_dir,
        active_page='eta_master',
    )


@master_bp.route('/eta/upload', methods=['POST'], endpoint='eta_master_upload_file')
@limiter.limit('10 per minute')
@require_admin
def eta_master_upload_file():
    if 'file' not in request.files or not request.files['file'].filename:
        flash('No file selected. Please choose an Excel file.', 'danger')
        return redirect(url_for('master.eta_master_upload'))

    file = request.files['file']
    if not file.filename.lower().endswith('.xlsx'):
        flash('Only .xlsx files are allowed.', 'danger')
        return redirect(url_for('master.eta_master_upload'))

    try:
        workbook = load_workbook(file, data_only=True)
        sheet = workbook.active

        if sheet.max_row < 2:
            flash('Excel file is empty or has no data rows.', 'danger')
            return redirect(url_for('master.eta_master_upload'))

        header_row = tuple(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        if not header_row or not any(header_row):
            flash('Excel file has no headers.', 'danger')
            return redirect(url_for('master.eta_master_upload'))

        header_map = _map_headers(header_row[0])
        required_cols = {alias: col for col, alias in header_map.items() if alias in REQUIRED_FIELDS}

        if set(required_cols.values()) != set(REQUIRED_FIELDS):
            flash(f'Missing required columns: {set(REQUIRED_FIELDS) - set(required_cols.values())}', 'danger')
            return redirect(url_for('master.eta_master_upload'))

        rows_data = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue
            row_dict = {}
            for col_idx, alias in header_map.items():
                if alias:
                    row_dict[alias] = row[col_idx] if col_idx < len(row) else None
            rows_data.append((row_idx, row_dict))

        inserted, updated, skipped, errors = _upsert_records(rows_data, file.filename)
        if errors:
            error_summary = '; '.join([f"Row {e['row']}: {e['error']}" for e in errors[:5]])
            flash(f'Inserted: {inserted}, Updated: {updated}, Errors: {len(errors)}. {error_summary}...', 'warning')
        else:
            flash(f'Successfully imported {inserted} new records and updated {updated} existing records.', 'success')

        return redirect(url_for('master.eta_master_upload'))
    except Exception as e:
        logger.exception('Error during ETA master file upload')
        flash(f'Error processing file: {str(e)}', 'danger')
        return redirect(url_for('master.eta_master_upload'))
